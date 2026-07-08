#!/usr/bin/env python3
"""
import_product_master.py — pre-processor for the StoreOps per-type
ProductMaster bulk import.

Reads the human-maintained master spreadsheets:
  * references/master_beer_sheet.xlsx       — beer (LCBO case pricing)
  * references/master_cigarette_sheet.xlsx  — cigarettes (carton → pack)
  * YV Vape Master Price List.xlsx          — vape (attributes + prices)

and writes one staging CSV per type, each matching the column order of the
matching _pm_<type>_staging tab (see ProductMaster.stagingLayout / Setup.js):

  scripts/product_master_beer_staging.csv
  scripts/product_master_cig_staging.csv
  scripts/product_master_vape_staging.csv
  scripts/product_master_other_staging.csv   (header-only template)

Pricing is DERIVED IN CODE on the server from the staged inputs, so the
resolved cost/sell/margin are left blank here.

WORKFLOW:
  1. python scripts/import_product_master.py
  2. In the spreadsheet, un-hide the _pm_<type>_staging tab.
  3. Select cell A3, paste the matching CSV (Sheets auto-parses commas).
  4. Run "📦 Import <Type> (from staging)" from the StoreOps menu.
  5. Read the inserted / updated / skipped / errors summary alert.

DEPENDENCY: openpyxl  (pip install openpyxl)

This script is LOCAL-ONLY — not pushed by CLASP (see .claspignore).
"""

import csv
import sys
from pathlib import Path

try:
    import openpyxl
except ImportError:
    sys.stderr.write(
        "ERROR: openpyxl not installed.\n"
        "Install it:  pip install openpyxl\n"
    )
    sys.exit(1)

# ────────────────────────────────────────────────────────────
# Paths
# ────────────────────────────────────────────────────────────
HERE = Path(__file__).resolve().parent
PROJECT_ROOT = HERE.parent
REFS = PROJECT_ROOT / "references"

BEER_FILE = REFS / "master_beer_sheet.xlsx"
CIG_FILE = REFS / "master_cigarette_sheet.xlsx"
VAPE_FILE = PROJECT_ROOT / "YV Vape Master Price List.xlsx"

OUT_BEER = HERE / "product_master_beer_staging.csv"
OUT_CIG = HERE / "product_master_cig_staging.csv"
OUT_VAPE = HERE / "product_master_vape_staging.csv"
OUT_OTHER = HERE / "product_master_other_staging.csv"

# ────────────────────────────────────────────────────────────
# Column layouts — MUST match ProductMaster.stagingLayout(type) in
# ProductMaster.js (CORE_STAGING_BASE + the type's input columns).
# ────────────────────────────────────────────────────────────
CORE_BASE = [
    "sku", "barcode", "product_name", "brand", "category", "subcategory",
    "pack_size", "unit", "supplier", "min_sell_price", "notes", "source_file",
]
BEER_HEADERS = CORE_BASE + [
    "sell_unit", "lcbo_case_price", "units_per_case", "packs_per_case",
    "cost_input", "deposit_per_unit", "target_margin_pct",
]
CIG_HEADERS = CORE_BASE + [
    "cost_per_carton", "packs_per_carton", "target_margin_pct", "review_sold",
]
VAPE_HEADERS = CORE_BASE + [
    "product_line", "form_factor", "puffs", "eliquid_ml", "nicotine", "flavor",
    "purchase_price", "sale_price", "sale_price_credit", "last_invoice_no",
]
OTHER_HEADERS = CORE_BASE + ["cost_price", "sell_price", "sell_price_credit"]


# ────────────────────────────────────────────────────────────
# Helpers
# ────────────────────────────────────────────────────────────
def s(v):
    """Stringify a cell value, trimming whitespace; '' for None."""
    return "" if v is None else str(v).strip()


def num(v):
    """Numeric coercion; '' for blanks/unparseable so they stay empty in CSV."""
    if v is None or v == "":
        return ""
    try:
        return float(v)
    except (TypeError, ValueError):
        return ""


def parse_money(v):
    """Numeric coercion that tolerates '$24.00' / '1,234.50'. '' for blanks."""
    if v is None or v == "":
        return ""
    if isinstance(v, (int, float)):
        return float(v)
    txt = str(v).replace("$", "").replace(",", "").strip()
    if not txt:
        return ""
    try:
        return float(txt)
    except (TypeError, ValueError):
        return ""


def cell(ws, row, col):
    """1-indexed sheet cell value."""
    return ws.cell(row=row, column=col).value


def empty_row(headers):
    return {h: "" for h in headers}


def open_first_sheet(path, prefer=None):
    wb = openpyxl.load_workbook(path, data_only=True)
    if prefer and prefer in wb.sheetnames:
        return wb[prefer]
    return wb.worksheets[0]


# ────────────────────────────────────────────────────────────
# Parser: Beer master (references/master_beer_sheet.xlsx)
# Columns (1-indexed): Product | Sell Unit | Category | SKU |
#   LCBO Case Price | Units/Case | Packs/Case | Cost | Deposit |
#   Margin % | Sell Price | Profit | Store Update
# Each (Product, Sell Unit) is its own row → its own core product.
#
# The sheet also tracks cooler/RTD, wine and cider — all of which use the
# SAME LCBO case-price + deposit pricing model as beer. They all map to
# category 'beer' (the pricing-model discriminator) with their real class
# kept in subcategory.
# ────────────────────────────────────────────────────────────
BEER_LCBO_CATS = {"beer", "cooler/rtd", "rtd", "cooler", "coolers", "wine", "cider", "seltzer"}


def parse_beer(path):
    if not path.exists():
        sys.stderr.write(f"WARN: {path.name} not found — skipping beer parser.\n")
        return []
    ws = open_first_sheet(path, "Master Beer Sheet")
    out = []
    for r in range(1, ws.max_row + 1):
        cat = s(cell(ws, r, 3)).lower()
        if cat not in BEER_LCBO_CATS:   # skips title / subtitle / header / blanks
            continue
        product = s(cell(ws, r, 1))
        if not product:
            continue
        row = empty_row(BEER_HEADERS)
        row["sku"]               = s(cell(ws, r, 4))
        row["product_name"]      = product
        row["category"]          = "beer"
        row["subcategory"]       = s(cell(ws, r, 3))   # Beer / Cooler-RTD / Wine / Cider
        row["unit"]              = s(cell(ws, r, 2))   # sell unit (Single / 6-Pack)
        row["source_file"]       = "master_beer_sheet.xlsx"
        row["sell_unit"]         = s(cell(ws, r, 2))
        row["lcbo_case_price"]   = parse_money(cell(ws, r, 5))
        row["units_per_case"]    = num(cell(ws, r, 6))
        row["packs_per_case"]    = parse_money(cell(ws, r, 7))
        row["cost_input"]        = parse_money(cell(ws, r, 8))
        row["deposit_per_unit"]  = parse_money(cell(ws, r, 9))
        row["target_margin_pct"] = num(cell(ws, r, 10))   # already 0..1
        out.append(row)
    return out


# ────────────────────────────────────────────────────────────
# Parser: Cigarette master (references/master_cigarette_sheet.xlsx)
# Columns (1-indexed): Supplier | Product | Category | Pack Size |
#   Cost/Carton | Packs/Carton | Cost/Pack | Margin % | Sell/Pack |
#   Markup % | Profit/Pack | Review (sold?)
# No SKU in this sheet → server dedups on brand + name.
# ────────────────────────────────────────────────────────────
def parse_cigarettes(path):
    if not path.exists():
        sys.stderr.write(f"WARN: {path.name} not found — skipping cigarette parser.\n")
        return []
    ws = open_first_sheet(path, "Master Cigarette Sheet")
    out = []
    for r in range(1, ws.max_row + 1):
        if s(cell(ws, r, 3)).lower() != "cigarettes":
            continue
        product = s(cell(ws, r, 2))
        if not product:
            continue
        row = empty_row(CIG_HEADERS)
        row["product_name"]      = product
        row["category"]          = "cigarettes"
        row["pack_size"]         = s(cell(ws, r, 4))    # 20s / 25s
        row["unit"]              = "pack"
        row["supplier"]          = s(cell(ws, r, 1))
        row["source_file"]       = "master_cigarette_sheet.xlsx"
        row["cost_per_carton"]   = parse_money(cell(ws, r, 5))
        row["packs_per_carton"]  = num(cell(ws, r, 6))
        row["target_margin_pct"] = num(cell(ws, r, 8))   # already 0..1
        row["review_sold"]       = s(cell(ws, r, 12))
        out.append(row)
    return out


# ────────────────────────────────────────────────────────────
# Parser: YV Vape Master Price List (vape)
# Columns (1-indexed): SKU | Brand | Product Line | Category | Form Factor |
#   Puffs | Pack Size | E-Liquid (mL) | Nicotine | Flavor | Purchase Price |
#   Sale Price | Sale Price (Credit) | Margin $ | Margin % | Supplier |
#   Last Invoice # | Notes
# Attributes get REAL columns (no longer crammed into product_name).
# ────────────────────────────────────────────────────────────
def parse_vape(path):
    if not path.exists():
        sys.stderr.write(f"WARN: {path.name} not found — skipping vape parser.\n")
        return []
    wb = openpyxl.load_workbook(path, data_only=True)
    if "Master Price List" not in wb.sheetnames:
        raise RuntimeError(
            f"{path.name}: 'Master Price List' tab not found (found: {wb.sheetnames})"
        )
    ws = wb["Master Price List"]
    if "sku" not in s(cell(ws, 1, 1)).lower():
        raise RuntimeError(f"{path.name}: expected SKU at A1. File layout shifted?")

    out = []
    for r in range(2, ws.max_row + 1):
        sku          = s(cell(ws, r, 1))
        brand        = s(cell(ws, r, 2))
        product_line = s(cell(ws, r, 3))
        form_factor  = s(cell(ws, r, 5))
        pack_size    = s(cell(ws, r, 7))
        nicotine     = s(cell(ws, r, 9))
        flavor       = s(cell(ws, r, 10))
        supplier     = s(cell(ws, r, 16))
        notes        = s(cell(ws, r, 18))

        if not sku and not brand and not product_line:
            continue   # end-of-data / blank grid row

        # Display name from brand + line + flavor (attributes live in detail).
        name = " ".join([p for p in [brand, product_line, flavor] if p]).strip() or sku

        invoice_raw = cell(ws, r, 17)
        last_invoice_no = ""
        if invoice_raw not in (None, ""):
            try:
                last_invoice_no = str(int(invoice_raw))
            except (TypeError, ValueError):
                last_invoice_no = str(invoice_raw).strip()

        row = empty_row(VAPE_HEADERS)
        row["sku"]               = sku
        row["product_name"]      = name
        row["brand"]             = brand
        row["category"]          = "vape"
        row["subcategory"]       = form_factor
        row["pack_size"]         = pack_size
        row["unit"]              = "each"
        row["supplier"]          = supplier
        row["notes"]             = notes
        row["source_file"]       = "YV Vape Master Price List.xlsx"
        row["product_line"]      = product_line
        row["form_factor"]       = form_factor
        row["puffs"]             = num(cell(ws, r, 6))
        row["eliquid_ml"]        = num(cell(ws, r, 8))
        row["nicotine"]          = nicotine
        row["flavor"]            = flavor
        row["purchase_price"]    = parse_money(cell(ws, r, 11))
        row["sale_price"]        = parse_money(cell(ws, r, 12))
        row["sale_price_credit"] = parse_money(cell(ws, r, 13))
        row["last_invoice_no"]   = last_invoice_no
        out.append(row)
    return out


# ────────────────────────────────────────────────────────────
# Dedup (mirrors the server's findDuplicate_)
# ────────────────────────────────────────────────────────────
def dedup(rows, key_fn):
    seen = set()
    final = []
    skipped = 0
    for r in rows:
        k = key_fn(r)
        if k in seen:
            skipped += 1
            continue
        seen.add(k)
        final.append(r)
    return final, skipped


def beer_key(r):
    sku = (r.get("sku") or "").strip().lower()
    unit = (r.get("unit") or "").strip().lower()
    if sku:
        return "sku:" + sku + "|" + unit
    return "bn:" + (r.get("brand") or "").strip().lower() + "|" + (r.get("product_name") or "").strip().lower() + "|" + unit


def sku_or_name_key(r):
    sku = (r.get("sku") or "").strip().lower()
    if sku:
        return "sku:" + sku
    return "bn:" + (r.get("brand") or "").strip().lower() + "|" + (r.get("product_name") or "").strip().lower()


# ────────────────────────────────────────────────────────────
# CSV writer
# ────────────────────────────────────────────────────────────
def write_csv(rows, headers, path):
    with path.open("w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=headers)
        w.writeheader()
        for r in rows:
            w.writerow(r)


# ────────────────────────────────────────────────────────────
# Main
# ────────────────────────────────────────────────────────────
def main():
    beer = parse_beer(BEER_FILE)
    beer, beer_dups = dedup(beer, beer_key)
    write_csv(beer, BEER_HEADERS, OUT_BEER)
    print(f"  Beer:        {len(beer):>4} rows  ({beer_dups} dup)  → {OUT_BEER.name}")

    cig = parse_cigarettes(CIG_FILE)
    cig, cig_dups = dedup(cig, sku_or_name_key)
    write_csv(cig, CIG_HEADERS, OUT_CIG)
    print(f"  Cigarettes:  {len(cig):>4} rows  ({cig_dups} dup)  → {OUT_CIG.name}")

    vape = parse_vape(VAPE_FILE)
    vape, vape_dups = dedup(vape, sku_or_name_key)
    write_csv(vape, VAPE_HEADERS, OUT_VAPE)
    print(f"  Vape:        {len(vape):>4} rows  ({vape_dups} dup)  → {OUT_VAPE.name}")

    # Grocery/other has no master source — emit a header-only template.
    write_csv([], OTHER_HEADERS, OUT_OTHER)
    print(f"  Other:       template (header only)  → {OUT_OTHER.name}")

    print("\n✅  Done. For each type:")
    print("    1. Un-hide the _pm_<type>_staging tab")
    print("    2. Paste the matching CSV at A3")
    print('    3. Run "📦 Import <Type> (from staging)" from the StoreOps menu')


if __name__ == "__main__":
    main()
